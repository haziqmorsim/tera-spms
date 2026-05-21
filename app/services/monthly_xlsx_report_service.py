from __future__ import annotations
import calendar
import math
import tempfile
from pathlib import Path
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from app.services.tnb_bill_parser import MONTH_LABEL_EN, ParsedTnbBill
from app.utils.time_utils import now_gmt8

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

def _safe_file_name(value: str) -> str:
    bad = '<>:"/\\|?*'
    for ch in bad:
        value = value.replace(ch, "_")
    return value.strip()

def _get_versioned_destination(dest_dir: Path, filename: str) -> Path:
    dest_dir.mkdir(parents=True, exist_ok=True)
    candidate = dest_dir / filename

    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    idx = 2

    while True:
        test = dest_dir / f"{stem}_v{idx}{suffix}"
        if not test.exists():
            return test
        idx += 1

def _fmt_num(value):
    return None if value is None else float(value)

def _sum_column(values: list[float | int | None]) -> float | None:
    nums = [float(v) for v in values if v is not None]
    return sum(nums) if nums else None

def _build_chart_image(
    parsed: ParsedTnbBill,
    months: list[str],
    values: list[float],
) -> Path:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
        chart_path = Path(tmp.name)

    fig, ax = plt.subplots(figsize=(10, 4.8))

    bars = ax.bar(months, values, width=0.6)

    ax.set_title(
        f"Company Name: {parsed.company_name or '-'}",
        pad=12,
        fontweight="bold",
    )
    ax.set_ylabel("RM", labelpad=10)
    ax.set_xlabel("Month", labelpad=10)

    y_max, major_unit = _get_dynamic_axis_scale(values)

    ax.set_ylim(0, y_max)

    ticks = list(range(0, int(y_max) + int(major_unit), int(major_unit)))
    ax.set_yticks(ticks)
    ax.set_yticklabels([f"{tick:,.0f}" for tick in ticks])

    ax.grid(axis="y", linestyle="-", alpha=0.35)
    ax.set_axisbelow(True)

    for bar, value in zip(bars, values):
        label = "0" if value == 0 else f"{value:,.2f}".rstrip("0").rstrip(".")
        label_y = bar.get_height() + (y_max * 0.015)

        ax.text(
            bar.get_x() + bar.get_width() / 2,
            label_y,
            label,
            ha="center",
            va="bottom",
            fontsize=8.5,
            fontweight="bold",
        )

    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)

    plt.tight_layout()
    fig.savefig(chart_path, dpi=180, bbox_inches="tight")
    plt.close(fig)

    return chart_path

def _get_dynamic_axis_scale(values: list[float]) -> tuple[int, int]:
    positive_values = [float(v) for v in values if v is not None and float(v) > 0]

    if not positive_values:
        return 1000, 100

    max_value = max(positive_values)

    if max_value <= 1000:
        major_unit = 200
        y_max = 2000
    elif max_value <= 5000:
        major_unit = 500
        y_max = 5000
    elif max_value <= 10000:
        major_unit = 1000
        y_max = 10000
    elif max_value <= 100000:
        major_unit = 10000
        y_max = 100000
    elif max_value <= 500000:
        major_unit = 50000
        y_max = int(math.ceil(max_value / major_unit) * major_unit)
    elif max_value <= 1000000:
        major_unit = 100000
        y_max = int(math.ceil(max_value / major_unit) * major_unit)
    elif max_value <= 10000000:
        major_unit = 1000000
        y_max = int(math.ceil(max_value / major_unit) * major_unit)
    else:
        major_unit = 5000000
        y_max = int(math.ceil(max_value / major_unit) * major_unit)

    if y_max <= max_value:
        y_max += major_unit

    return int(y_max), int(major_unit)

def _apply_border_to_merged_range(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER

def build_monthly_xlsx_report(
    parsed: ParsedTnbBill,
    *,
    output_dir: str | Path = "reports/monthly",
) -> Path:
    output_dir = Path(output_dir)

    period_end = parsed.period_end or now_gmt8().date()
    suffix = period_end.strftime("%m-%Y")
    title_part = parsed.account_no or "tnb_bill"
    filename = _safe_file_name(f"monthly_report_{suffix}_{title_part}.xlsx")
    out_path = _get_versioned_destination(output_dir, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    widths = {
        "A": 14,
        "B": 6,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 14,
        "G": 14,
        "H": 10,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 11,
        "N": 14,
        "O": 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = "MONTHLY ELECTRIC BILL REPORT"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    meta = [
        ("Company", parsed.company_name),
        ("Account No", parsed.account_no),
        ("Invoice No", parsed.invoice_no),
        ("Tariff Description", parsed.tariff_desc),
        ("Bill Date", parsed.bill_date.strftime("%d-%m-%Y") if parsed.bill_date else None),
        (
            "Billing Period",
            f"{parsed.period_start.strftime('%d-%m-%Y')} to {parsed.period_end.strftime('%d-%m-%Y')}"
            if parsed.period_start and parsed.period_end
            else None,
        ),
        ("No. of Days", parsed.no_of_days),
        ("State", parsed.state),
        ("Address", parsed.address),
        ("Security Deposit (RM)", _fmt_num(parsed.security_deposit_rm)),
        ("Payment Amount (RM)", _fmt_num(parsed.payment_amount_rm)),
        ("Generated At", now_gmt8().strftime("%d-%m-%Y %H:%M:%S")),
    ]

    row = 3

    for label, value in meta:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.fill = SUBHEADER_FILL
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        _apply_border_to_merged_range(ws, row, 1, 2)

        for col in range(1, 3):
            ws.cell(row=row, column=col).fill = SUBHEADER_FILL

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
        value_cell = ws.cell(row=row, column=3, value=value)
        value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        _apply_border_to_merged_range(ws, row, 3, 8)

        row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    section_cell = ws.cell(row=row, column=1)
    section_cell.value = f"TNB Meter Acc. No.: {parsed.account_no or '-'}"
    section_cell.font = Font(bold=True, color="FFFFFF")
    section_cell.fill = HEADER_FILL
    section_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18

    row += 1

    headers = [
        "Month",
        "No. of\nDays",
        "Total TNB\nMonthly Bill\n(RM)",
        "Total Usage\n(kWh)",
        "Max\nDemand\n(kW)",
        "Peak Usage\n(kWh)",
        "Off-Peak Usage\n(kWh)",
        "AFA (RM)",
        "Peak Tariff\n(RM)",
        "Off-Peak Tariff\n(RM)",
        "Capacity Charge\n(RM)",
        "Network Charge\n(RM)",
        "Retail Charge\n(RM)",
        "Current Usage\nCharge\n(RM)",
        "KWTBB (RM)",
    ]

    header_row = row

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 42

    row += 1
    data_start_row = row

    history_map = {(item.year, item.month): item for item in parsed.history}
    current_year = period_end.year

    for month_no in range(1, 13):
        month_label = MONTH_LABEL_EN[month_no]
        hist = history_map.get((current_year, month_no))

        is_current_month = (
            parsed.period_end is not None
            and parsed.period_end.year == current_year
            and parsed.period_end.month == month_no
        )

        afa_value = None
        if is_current_month:
            afa_value = parsed.afa_rm if parsed.afa_rm is not None else 0.0

        values = [
            month_label,
            parsed.no_of_days
            if is_current_month and parsed.no_of_days is not None
            else calendar.monthrange(current_year, month_no)[1],
            _fmt_num(hist.total_tnb_monthly_bill_rm) if hist else None,
            _fmt_num(hist.total_usage_kwh) if hist else None,
            _fmt_num(parsed.max_demand_kw) if is_current_month else None,
            _fmt_num(parsed.peak_usage_kwh) if is_current_month else None,
            _fmt_num(parsed.offpeak_usage_kwh) if is_current_month else None,
            _fmt_num(afa_value),
            _fmt_num(parsed.peak_tariff_rm) if is_current_month else None,
            _fmt_num(parsed.offpeak_tariff_rm) if is_current_month else None,
            _fmt_num(parsed.capacity_charge_rm) if is_current_month else None,
            _fmt_num(parsed.network_charge_rm) if is_current_month else None,
            _fmt_num(parsed.retail_charge_rm) if is_current_month else None,
            _fmt_num(parsed.current_month_charge_rm) if is_current_month else None,
            _fmt_num(parsed.kwtbb_rm) if is_current_month else None,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 18
        row += 1

    total_row = row

    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(2, 16):
        values = [ws.cell(r, col_idx).value for r in range(data_start_row, total_row)]
        cell = ws.cell(row=total_row, column=col_idx, value=_sum_column(values))
        cell.font = Font(bold=True)
        cell.fill = SUBHEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[total_row].height = 18

    for r in range(data_start_row, total_row + 1):
        ws.cell(r, 2).number_format = "0"
        for c in range(3, 16):
            ws.cell(r, c).number_format = "#,##0.00"

    chart_months = [MONTH_LABEL_EN[m] for m in range(1, 13)]
    chart_values: list[float] = []

    for month_no in range(1, 13):
        hist = history_map.get((current_year, month_no))
        if hist and hist.total_tnb_monthly_bill_rm is not None:
            chart_values.append(float(hist.total_tnb_monthly_bill_rm))
        else:
            chart_values.append(0.0)

    chart_path = _build_chart_image(parsed, chart_months, chart_values)
    img = XLImage(str(chart_path))
    img.width = 760
    img.height = 360
    ws.add_image(img, f"D{total_row + 3}")

    wb.save(out_path)
    return out_path