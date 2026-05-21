from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1EA5FF")
SEVERITY_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="E44545"),
    "MAJOR": PatternFill("solid", fgColor="FF8000"),
    "MINOR": PatternFill("solid", fgColor="FFBB33"),
    "WARNING": PatternFill("solid", fgColor="4FAFF5"),
    "INFO": PatternFill("solid", fgColor="999999"),
}
THIN_SIDE = Side(style="thin", color="000000")
ALL_BORDERS = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def _normalize_text(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def _find_city_column(df: pd.DataFrame) -> str | None:
    for col in df.columns:
        col_norm = str(col).strip().lower()
        if col_norm == "city":
            return col
    return None


def _find_plant_name_column(df: pd.DataFrame) -> str | None:
    candidates = [
        "plant_name",
        "plant name",
        "name",
    ]
    normalized = {str(col).strip().lower(): col for col in df.columns}

    for candidate in candidates:
        if candidate in normalized:
            return normalized[candidate]

    return None


def load_plant_city_map(excel_path: str | Path) -> dict[str, str]:
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Plant information workbook not found: {excel_path}")

    df = pd.read_excel(excel_path)

    plant_col = _find_plant_name_column(df)
    city_col = _find_city_column(df)

    if not plant_col or not city_col:
        raise ValueError(
            "Could not find plant name / city columns in the plant information workbook."
        )

    mapping: dict[str, str] = {}
    for _, row in df.iterrows():
        plant_name = row.get(plant_col)
        city = row.get(city_col)

        if pd.isna(plant_name):
            continue

        plant_key = _normalize_text(plant_name)
        city_value = "" if pd.isna(city) else _normalize_text(city)

        if plant_key:
            mapping[plant_key] = city_value

    return mapping


def build_alarm_xlsx_report(
    alarm_rows: list[dict],
    plant_info_excel_path: str | Path,
    output_path: str | Path,
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    plant_city_map = load_plant_city_map(plant_info_excel_path)

    prepared_rows: list[dict] = []
    for row in alarm_rows:
        plant_name = _normalize_text(row.get("plant_name"))
        city = plant_city_map.get(plant_name, "")

        prepared_rows.append(
            {
                "PLANT NAME": plant_name,
                "CITY": city,
                "DEVICE NAME": _normalize_text(row.get("device_name")),
                "DEVICE SN": _normalize_text(row.get("device_sn")),
                "ALARM ID": _normalize_text(row.get("alarm_id")),
                "ALARM NAME": _normalize_text(row.get("alarm_name")),
                "SEVERITY": _normalize_text(row.get("severity")),
            }
        )

    prepared_rows.sort(
        key=lambda x: (
            x.get("CITY", ""),
            x.get("PLANT NAME", ""),
            x.get("DEVICE NAME", ""),
            x.get("ALARM NAME", ""),
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "ALARMS_REPORT"

    headers = [
        "PLANT NAME",
        "CITY",
        "DEVICE NAME",
        "DEVICE SN",
        "ALARM ID",
        "ALARM NAME",
        "SEVERITY",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = ALL_BORDERS

    for row_idx, item in enumerate(prepared_rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = item.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = ALL_BORDERS
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if header == "SEVERITY":
                fill = SEVERITY_FILLS.get(str(value).upper())
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)

        ws.row_dimensions[row_idx].height = 24

    ws.row_dimensions[1].height = 24

    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            value_len = len(str(value)) if value is not None else 0
            if value_len > max_len:
                max_len = value_len

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.border = ALL_BORDERS

    wb.save(output_path)
    return str(output_path)