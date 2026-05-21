from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1EA5FF")
STATUS_FILLS = {
    "NORMAL": PatternFill("solid", fgColor="41BA41"),
    "OFFLINE": PatternFill("solid", fgColor="999999"), 
    "STANDBY": PatternFill("solid", fgColor="FFBB33"), 
    "FAULTY": PatternFill("solid", fgColor="E44545"),
}
THIN_SIDE = Side(style="thin", color="000000")
ALL_BORDERS = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

def _to_upper_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.upper()
    return value

def _normalize_status(status, psh_value):
    status_text = str(status or "").strip().upper()

    try:
        psh_num = float(psh_value) if psh_value is not None else None
    except Exception:
        psh_num = None

    if status_text == "NORMAL" and psh_num == 0:
        return "OFFLINE"
    
    return status_text

def _safe_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None
    
def build_low_psh_xlsx_report(report_df, output_path: str | Path) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "low_psh_plants_report"

    desired_order = [
        "plant_name", 
        "status", 
        "country", 
        "city", 
        "grid_connection_date", 
        "total_string_capacity_kwp", 
        "current_power_kw", 
        "specific_energy_kwh_kwp", 
        "yield_today_kwh", 
        "total_yield_kwh", 
        "psh", 
        "city_plant_count", 
        "city_avg_psh", 
        "threshold_psh", 
        "performance_status", 
        "psh_deviation_pct_vs_city_avg", 
        "underperforming",
    ]

    existing_cols = [col for col in desired_order if col in report_df.columns]
    remaining_cols = [col for col in report_df.columns if col not in existing_cols]
    ordered_cols = existing_cols + remaining_cols

    for col_idx, col_name in enumerate(ordered_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = ALL_BORDERS

    for row_idx, (_, row) in enumerate(report_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(ordered_cols, start=1):
            value = row[col_name] if col_name in row else None

            if col_name == "status":
                value = _normalize_status(value, row.get("psh"))
            elif isinstance(value, str):
                value = value.upper()
            elif col_name == "underperforming":
                value = str(value).upper()

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = ALL_BORDERS
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_name == "status":
                status_fill = STATUS_FILLS.get(str(value or "").upper())
                if status_fill:
                    cell.fill = status_fill
                    cell.font = Font(bold=True)

        ws.row_dimensions[row_idx].height = 24

    ws.row_dimensions[1].height = 24
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(ordered_cols, start=1):
        max_len = len(str(col_name).upper())
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            val_len = len(str(val)) if val is not None else 0
            if val_len > max_len:
                max_len = val_len

        adjusted_width = min(max_len + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = ALL_BORDERS

    wb.save(output_path)
    return str(output_path)