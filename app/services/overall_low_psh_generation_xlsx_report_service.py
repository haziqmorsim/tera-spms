from __future__ import annotations
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session
from app.services.report_path_service import EXCEL_OVERALL_DIR
from app.utils.time_utils import today_gmt8

HEADER_FILL = PatternFill("solid", fgColor="6EC1E4")
NORMAL_FILL = PatternFill("solid", fgColor="41BA41")
OFFLINE_FILL = PatternFill("solid", fgColor="999999")
STANDBY_FILL = PatternFill("solid", fgColor="FFBB33")
FAULTY_FILL = PatternFill("solid", fgColor="FF0000")
UNKNOWN_FILL = PatternFill("solid", fgColor="D9D9D9")

THIN_SIDE = Side(style="thin", color="000000")
ALL_BORDERS = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

def _safe_upper(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip().upper()

def _safe_float(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None

        return float(value)

    except Exception:
        return None

def _format_date(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return str(value).strip()

def _status_fill(status: str) -> PatternFill:
    status_upper = _safe_upper(status)

    if status_upper == "NORMAL":
        return NORMAL_FILL

    if status_upper == "OFFLINE":
        return OFFLINE_FILL

    if status_upper == "STANDBY":
        return STANDBY_FILL

    if status_upper == "FAULTY":
        return FAULTY_FILL

    return UNKNOWN_FILL

def _format_string_number(value: Any) -> str:
    text_value = str(value or "").strip()

    if not text_value:
        return ""

    match = re.search(
        r"PV\s*(\d+)\s+input\s+current\s*\(\s*A\s*\)",
        text_value,
        re.I,
    )

    if match:
        return f"PV {int(match.group(1))} Current (A) String"

    return text_value

def resolve_low_psh_generation_report_day(db: Session) -> date | None:
    try:
        latest_string_report_day = (
            db.execute(
                text(
                    """
                    SELECT report_day
                    FROM generated_reports
                    WHERE upper(report_type) = 'LOW_PERFORMING_STRING_XLSX'
                      AND report_day IS NOT NULL
                    ORDER BY generated_at DESC NULLS LAST, report_day DESC
                    LIMIT 1
                    """
                )
            )
            .scalar_one_or_none()
        )

        if latest_string_report_day is not None:
            return latest_string_report_day

    except Exception:
        db.rollback()

    try:
        latest_string_data_day = (
            db.execute(
                text(
                    """
                    SELECT MAX(run_day)
                    FROM low_performing_strings_latest
                    """
                )
            )
            .scalar_one_or_none()
        )

        if latest_string_data_day is not None:
            return latest_string_data_day

    except Exception:
        db.rollback()

    try:
        latest_inverter_day = (
            db.execute(
                text(
                    """
                    SELECT MAX(run_day)
                    FROM low_performing_inverters_latest
                    """
                )
            )
            .scalar_one_or_none()
        )

        if latest_inverter_day is not None:
            return latest_inverter_day

    except Exception:
        db.rollback()

    try:
        latest_plant_day = (
            db.execute(
                text(
                    """
                    SELECT MAX(run_day)
                    FROM low_psh_plants_by_city_latest
                    """
                )
            )
            .scalar_one_or_none()
        )

        if latest_plant_day is not None:
            return latest_plant_day

    except Exception:
        db.rollback()

    return None

def fetch_overall_low_psh_generation_rows(
    db: Session,
    *,
    report_day: date | None = None,
) -> list[dict]:
    if report_day is None:
        report_day = resolve_low_psh_generation_report_day(db)

    if report_day is None:
        return []

    rows = (
        db.execute(
            text(
                """
                WITH latest_plant_rows AS (
                    SELECT DISTINCT ON (upper(trim(l.plant_name)))
                        l.plant_name,
                        COALESCE(
                            NULLIF(l.plant_status, ''),
                            NULLIF(l.performance_status, ''),
                            NULLIF(p.status, ''),
                            'UNKNOWN'
                        ) AS status,
                        COALESCE(
                            l.total_string_capacity_kwp,
                            p.total_string_capacity_kwp
                        ) AS total_capacity_kwp,
                        COALESCE(
                            l.grid_connection_date,
                            p.grid_connection_date
                        ) AS grid_connected_date
                    FROM low_psh_plants_by_city_latest l
                    LEFT JOIN plants p
                        ON upper(trim(p.name)) = upper(trim(l.plant_name))
                    WHERE l.run_day <= :report_day
                    ORDER BY upper(trim(l.plant_name)), l.run_day DESC
                ),
                latest_inverter_rows AS (
                    SELECT DISTINCT ON (
                        upper(trim(plant_name)),
                        upper(trim(inverter_name))
                    )
                        plant_name,
                        inverter_name,
                        inverter_sn,
                        plant_status
                    FROM low_performing_inverters_latest
                    WHERE run_day <= :report_day
                      AND underperforming = true
                    ORDER BY
                        upper(trim(plant_name)),
                        upper(trim(inverter_name)),
                        run_day DESC
                )
                SELECT
                    s.plant_name,
                    COALESCE(
                        NULLIF(s.plant_status, ''),
                        NULLIF(i.plant_status, ''),
                        NULLIF(p.status, ''),
                        'UNKNOWN'
                    ) AS status,
                    p.total_capacity_kwp,
                    p.grid_connected_date,
                    s.inverter_name,
                    COALESCE(
                        NULLIF(s.inverter_sn, ''),
                        NULLIF(i.inverter_sn, '')
                    ) AS inverter_sn,
                    s.string_name
                FROM low_performing_strings_latest s
                LEFT JOIN latest_plant_rows p
                    ON upper(trim(p.plant_name)) = upper(trim(s.plant_name))
                LEFT JOIN latest_inverter_rows i
                    ON upper(trim(i.plant_name)) = upper(trim(s.plant_name))
                   AND upper(trim(i.inverter_name)) = upper(trim(s.inverter_name))
                WHERE s.run_day = :report_day
                  AND s.underperforming = true
                ORDER BY
                    upper(trim(s.plant_name)),
                    upper(trim(s.inverter_name)),
                    CASE
                        WHEN s.string_name ~* 'PV[ ]*[0-9]+'
                        THEN CAST(substring(s.string_name from 'PV[ ]*([0-9]+)') AS integer)
                        ELSE 9999
                    END,
                    upper(trim(s.string_name))
                """
            ),
            {"report_day": report_day},
        )
        .mappings()
        .all()
    )

    prepared: list[dict] = []

    for row in rows:
        prepared.append(
            {
                "plant_name": _safe_upper(row.get("plant_name")),
                "status": _safe_upper(row.get("status")),
                "total_capacity_kwp": _safe_float(row.get("total_capacity_kwp")),
                "grid_connected_date": _format_date(row.get("grid_connected_date")),
                "inverter_number": str(row.get("inverter_name") or "").strip(),
                "sn_inverter": str(row.get("inverter_sn") or "").strip(),
                "string_number": _format_string_number(row.get("string_name")),
            }
        )

    return prepared

def _merge_same_plant_cells(ws, start_row: int, end_row: int) -> None:
    if end_row <= start_row:
        return

    for col_idx in [1, 2, 3, 4]:
        ws.merge_cells(
            start_row=start_row,
            start_column=col_idx,
            end_row=end_row,
            end_column=col_idx,
        )

        cell = ws.cell(row=start_row, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ALL_BORDERS

def _merge_same_inverter_cells(ws, start_row: int, end_row: int) -> None:
    if end_row <= start_row:
        return

    for col_idx in [5, 6]:
        ws.merge_cells(
            start_row=start_row,
            start_column=col_idx,
            end_row=end_row,
            end_column=col_idx,
        )

        cell = ws.cell(row=start_row, column=col_idx)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ALL_BORDERS

def build_overall_low_psh_generation_xlsx_report(
    rows: list[dict],
    output_path: str | Path,
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "LOW_PSH_GENERATION"

    headers = [
        "PLANT NAME",
        "STATUS",
        "TOTAL CAPACITY (KWP)",
        "GRID CONNECTION DATE",
        "INVERTER NAME",
        "INVERTER SN",
        "STRING NUMBER",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ALL_BORDERS

    if not rows:
        ws.cell(row=2, column=1, value="NO LOW-PERFORMANCE GENERATION RECORDS DETECTED")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=1).border = ALL_BORDERS
        ws.row_dimensions[2].height = 28

    else:
        row_no = 2

        for item in rows:
            values = [
                item.get("plant_name"),
                item.get("status"),
                item.get("total_capacity_kwp"),
                item.get("grid_connected_date"),
                item.get("inverter_number"),
                item.get("sn_inverter"),
                item.get("string_number"),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_no, column=col_idx, value=value)
                cell.border = ALL_BORDERS
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if col_idx == 2:
                    cell.fill = _status_fill(str(value or ""))
                    cell.font = Font(color="000000")

            ws.row_dimensions[row_no].height = 22
            row_no += 1

        inverter_group_start = 2
        current_inverter_key = (
            ws.cell(row=2, column=1).value,
            ws.cell(row=2, column=5).value,
            ws.cell(row=2, column=6).value,
        )

        for row_idx in range(3, ws.max_row + 2):
            next_inverter_key = (
                ws.cell(row=row_idx, column=1).value if row_idx <= ws.max_row else None,
                ws.cell(row=row_idx, column=5).value if row_idx <= ws.max_row else None,
                ws.cell(row=row_idx, column=6).value if row_idx <= ws.max_row else None,
            )

            if next_inverter_key != current_inverter_key:
                _merge_same_inverter_cells(ws, inverter_group_start, row_idx - 1)
                inverter_group_start = row_idx
                current_inverter_key = next_inverter_key

        plant_group_start = 2
        current_plant = ws.cell(row=2, column=1).value

        for row_idx in range(3, ws.max_row + 2):
            next_plant = ws.cell(row=row_idx, column=1).value if row_idx <= ws.max_row else None

            if next_plant != current_plant:
                _merge_same_plant_cells(ws, plant_group_start, row_idx - 1)
                plant_group_start = row_idx
                current_plant = next_plant

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    widths = {
        1: 44,
        2: 14,
        3: 22,
        4: 22,
        5: 28,
        6: 24,
        7: 30,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row:
            cell.border = ALL_BORDERS

    wb.save(output_path)
    return str(output_path)

def build_latest_overall_low_psh_generation_xlsx_report(
    db: Session,
    *,
    report_day: date | None = None,
    output_path: str | Path | None = None,
) -> str:
    report_day = report_day or resolve_low_psh_generation_report_day(db) or today_gmt8()

    if output_path is None:
        output_path = (
            EXCEL_OVERALL_DIR
            / f"low_psh_generation_report_{report_day.strftime('%d-%m-%Y')}.xlsx"
        )

    rows = fetch_overall_low_psh_generation_rows(db, report_day=report_day)

    return build_overall_low_psh_generation_xlsx_report(
        rows=rows,
        output_path=output_path,
    )