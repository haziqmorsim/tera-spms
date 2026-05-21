from __future__ import annotations
from datetime import date, datetime
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F9FE5")
TRUE_FILL = PatternFill("solid", fgColor="FF4D4D")
FALSE_FILL = PatternFill("solid", fgColor="C6EFCE")
THIN_SIDE = Side(style="thin", color="000000")
ALL_BORDERS = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

def _get(row: dict, *keys: str, default: Any = None) -> Any:
    for key in keys:
        if key in row and row[key] is not None:
            return row[key]

        lower_key = key.lower()
        if lower_key in row and row[lower_key] is not None:
            return row[lower_key]

        upper_key = key.upper()
        if upper_key in row and row[upper_key] is not None:
            return row[upper_key]

    return default

def _format_date(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    return str(value)

def _format_number(value: Any, decimals: int = 3) -> float | None:
    if value is None or value == "":
        return None

    try:
        return round(float(value), decimals)
    except Exception:
        return None

def _format_bool(value: Any) -> str:
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    text = str(value or "").strip().lower()

    if text in {"true", "1", "yes", "y"}:
        return "TRUE"

    if text in {"false", "0", "no", "n"}:
        return "FALSE"

    return str(value or "")

def build_low_performing_string_xlsx_report(
    rows: list[dict],
    output_path: str | Path,
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "low_performing_strings"

    headers = [
        "RUN DAY",
        "PLANT NAME",
        "CITY",
        "PLANT STATUS",
        "INVERTER NAME",
        "INVERTER SN",
        "STRING NAME",
        "STRING TOTAL CURRENT",
        "BENCHMARK STRING CURRENT",
        "THRESHOLD STRING CURRENT",
        "DEVIATION PCT VS BENCHMARK",
        "UNDERPERFORMING",
        "REASON",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ALL_BORDERS

    if not rows:
        ws.cell(row=2, column=1, value="NO LOW-PERFORMING STRINGS DETECTED")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=1).border = ALL_BORDERS
    else:
        for row_idx, row in enumerate(rows, start=2):
            values = [
                _format_date(_get(row, "run_day")),
                _get(row, "plant_name"),
                _get(row, "city"),
                _get(row, "plant_status"),
                _get(row, "inverter_name"),
                _get(row, "inverter_sn"),
                _get(row, "string_name"),
                _format_number(_get(row, "string_total_current"), 3),
                _format_number(_get(row, "benchmark_string_current"), 3),
                _format_number(_get(row, "threshold_string_current"), 3),
                _format_number(_get(row, "deviation_pct_vs_benchmark"), 2),
                _format_bool(_get(row, "underperforming")),
                _get(row, "reason"),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = ALL_BORDERS
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if col_idx == 12:
                    if str(value).upper() == "TRUE":
                        cell.fill = TRUE_FILL
                        cell.font = Font(bold=True, color="000000")
                    elif str(value).upper() == "FALSE":
                        cell.fill = FALSE_FILL
                        cell.font = Font(bold=True, color="000000")

            ws.row_dimensions[row_idx].height = 28

    widths = {
        1: 14,
        2: 34,
        3: 18,
        4: 16,
        5: 28,
        6: 22,
        7: 24,
        8: 22,
        9: 24,
        10: 24,
        11: 26,
        12: 18,
        13: 48,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return str(output_path)