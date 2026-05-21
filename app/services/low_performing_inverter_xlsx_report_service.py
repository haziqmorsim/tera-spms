from __future__ import annotations
from pathlib import Path
from typing import Iterable
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1EA5FF")
LOW_FILL = PatternFill("solid", fgColor="E44545")
NORMAL_FILL = PatternFill("solid", fgColor="41BA41")
THIN_SIDE = Side(style="thin", color="000000")
ALL_BORDERS = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

def _safe_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None

def build_low_performing_inverter_xlsx_report(
    rows: Iterable[dict],
    output_path: str | Path,
) -> str:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "low_performing_inverters"

    columns = [
        "run_day",
        "plant_name",
        "city",
        "plant_status",
        "plant_psh",
        "city_avg_psh",
        "inverter_name",
        "inverter_sn",
        "inverter_psh",
        "benchmark_inverter_psh",
        "threshold_inverter_psh",
        "deviation_pct_vs_benchmark",
        "underperforming",
        "reason",
    ]

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ALL_BORDERS

    row_no = 2
    for row in rows:
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name)
            if col_name == "underperforming":
                value = str(bool(value)).upper()
            elif col_name in {
                "plant_psh",
                "city_avg_psh",
                "inverter_psh",
                "benchmark_inverter_psh",
                "threshold_inverter_psh",
                "deviation_pct_vs_benchmark",
            }:
                value = _safe_float(value)
            elif isinstance(value, str):
                value = value.upper()

            cell = ws.cell(row=row_no, column=col_idx, value=value)
            cell.border = ALL_BORDERS
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if col_name == "underperforming":
                cell.fill = LOW_FILL if value == "TRUE" else NORMAL_FILL
                cell.font = Font(bold=True)

        ws.row_dimensions[row_no].height = 30
        row_no += 1

    if row_no == 2:
        ws.cell(row=2, column=1, value="NO LOW-PERFORMING INVERTERS DETECTED")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=1).border = ALL_BORDERS

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name.upper())
        for r in range(2, ws.max_row + 1):
            value = ws.cell(row=r, column=col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    wb.save(output_path)
    return str(output_path)